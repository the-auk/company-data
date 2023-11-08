# Import the necessary libraries
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import openpyxl
from openpyxl.styles import Font
import random
import time

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 500

service = Service()
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

url = 'https://www.irs.gov/credits-deductions/residential-clean-energy-credit'

driver.get(url)

linksToFetch = {}
hyperLinks={}
def scrape_article():
    currentRow=1
    titleFont = Font(size=18, bold=True)
    subTitleFont = Font(size=15, bold=True)
    try:
        title = driver.find_element(By.XPATH, "//h1[contains(@class, 'title')]")
        article = driver.find_element(By.XPATH,"//article[contains(@class, 'pup-article')]")
        parent = article.find_element(By.XPATH, ".//div[contains(@class,'field')]")
        children = parent.find_elements(By.XPATH, "*")
        links = parent.find_elements(By.XPATH, ".//a")
        cell = sheet.cell(row=currentRow, column=1, value=title.text)
        cell.font = titleFont
        currentRow+=1
        for child in children:
            text = child.text
            cell = sheet.cell(row=currentRow, column=1, value=text)
            currentRow+=1
            tag = child.tag_name
            if tag=="h2":
                cell.font = subTitleFont
        cell = sheet.cell(row=currentRow, column=1, value="--------PAGE LINKS---------")
        currentRow+=1
        font = Font(color="FF000080")
        for link in links:
            temp = link.get_attribute("href")
            temp1 = temp.split("#")[0]
            if temp1 in linksToFetch or temp1 == url:
                continue
            if temp1.endswith(".pdf"):
                hyperLinks[temp]=1
            else:
                cell = sheet.cell(row=currentRow, column=1, value=temp1)
                cell.hyperlink = temp1
                cell.font = font
                currentRow+=1
                linksToFetch[temp1]=1

        if len(hyperLinks)>0:
            cell = sheet.cell(row=currentRow, column=1, value="--------PDF LINKS---------")
            currentRow+=1        


        for link in hyperLinks:
            cell = sheet.cell(row=currentRow, column=1, value=link)
            cell.hyperlink = link
            cell.font = font
            currentRow+=1
        cell = sheet.cell(row=currentRow, column=1, value="--------PAGE ENDS---------")
        currentRow+=1
        cell = sheet.cell(row=currentRow, column=1, value=" ")
        currentRow+=1
        for link in linksToFetch:
            print(link)
            rand = random.uniform(1.2, 2.67)
            time.sleep(rand)
            driver.get(link)
            try:
                children=[]
                try:
                    article = driver.find_element(By.XPATH,"//article[contains(@class, 'pup-article')]")
                    parent = article.find_element(By.XPATH, ".//div[contains(@class,'field')]")
                    children = parent.find_elements(By.XPATH, "*")
                    title = driver.find_element(By.XPATH, "//h1[contains(@class, 'title')]")
                    if len(children)==0:
                        raise Exception()
                    cell = sheet.cell(row=currentRow, column=1, value=title.text)
                    cell.font = titleFont
                    currentRow+=1
                except:
                    try:
                        title = driver.find_element(By.XPATH, "//h1[contains(@class, 'title')]")
                        parent1 = driver.find_element(By.XPATH, ".//div[@id='edit-content-top']")
                        parent2 = driver.find_element(By.XPATH, ".//div[@id='edit-content-bottom']")
                        child1 = parent1.find_elements(By.XPATH, "*")
                        child2 = parent2.find_elements(By.XPATH, "*")       
                        children = child1+child2
                        cell = sheet.cell(row=currentRow, column=1, value=title.text)
                        cell.font = titleFont
                        currentRow+=1
                    except:
                        continue

                for child in children:
                    text = child.text
                    cell = sheet.cell(row=currentRow, column=1, value=text)
                    currentRow+=1
                    tag = child.tag_name
                    if tag=="h2":
                        cell.font = subTitleFont
                cell = sheet.cell(row=currentRow, column=1, value="--------PAGE ENDS---------")
                currentRow+=1
                cell = sheet.cell(row=currentRow, column=1, value=" ")
                currentRow+=1
            except:
                continue
        
        workbook.save("residential.xlsx")


    except Exception as e:
        print(f"An error occurred: {str(e)}")

scrape_article()

# Close the WebDriver
driver.quit()
